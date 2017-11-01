import argparse
import glob
import os
import sys
import traceback

import numpy
import openpyxl

from . import las


class ExcelConverter(object):

    '''Provide ability to export LAS data into an Excel spreadsheet.

    Arguments:
        las (:class:`lasio.las.LASFile` object)

    '''

    def __init__(self, las):
        self.set_las(las)

    def set_las(self, las):
        '''Set LASFile object.

        Arguments:
            las (:class:`lasio.las.LASFile` object)

        '''
        self.las = las
        self.generate_workbook()
        return self

    def generate_workbook(self):
        '''Generate the Excel workbook object.

        Two sheets are created:

        * Header: contains all the header sections and metadata
        * Curves: contains the data 

        '''
        wb = openpyxl.Workbook()
        header = wb['Sheet']
        header.title = 'Header'
        curves = wb.create_sheet()
        curves.title = 'Curves'

        def write_cell(sh, i, j, value):
            c = sh.cell(row=i + 1, column=j + 1)
            c.value = value

        write_cell(header, 0, 0, 'Section')
        write_cell(header, 0, 1, 'Mnemonic')
        write_cell(header, 0, 2, 'Unit')
        write_cell(header, 0, 3, 'Value')
        write_cell(header, 0, 4, 'Description')

        sections = [
            ('~Version', self.las.version),
            ('~Well', self.las.well),
            ('~Parameter', self.las.params),
            ('~Curves', self.las.curves),
        ]

        n = 1
        for sect_name, sect in sections:
            for i, item in enumerate(sect.values()):
                write_cell(header, n, 0, sect_name)
                write_cell(header, n, 1, item.mnemonic)
                write_cell(header, n, 2, item.unit)
                write_cell(header, n, 3, item.value)
                write_cell(header, n, 4, item.descr)
                n += 1

        for i, curve in enumerate(self.las.curves):
            write_cell(curves, 0, i, curve.mnemonic)
            for j, value in enumerate(curve.data):
                if numpy.isnan(value):
                    write_cell(curves, j + 1, i, '')
                else:
                    write_cell(curves, j + 1, i, value)

        self.workbook = wb
        return self

    def write(self, xlsxfn):
        '''Write the Excel workbook to an ``.xlsx`` file.

        Arguments:
            xlsxfn (str): filename (will be overwritten without warning)

        '''
        assert xlsxfn.lower().endswith('.xlsx')

        self.workbook.save(xlsxfn)


def main():
    args = get_parser().parse_args(sys.argv[1:])
    lasfn = args.LAS_filename
    xlsxfn = args.XLSX_filename

    l = las.LASFile(lasfn)
    converter = ExcelConverter(l)
    converter.write(xlsxfn)


def get_parser():
    parser = argparse.ArgumentParser('Convert LAS file to XLSX', 
        formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument('LAS_filename')
    parser.add_argument('XLSX_filename')
    return parser


def main_bulk():
    args = get_bulk_parser().parse_args(sys.argv[1:])
    paths = []
    if args.recursive:
        for dirpath, dirnames, filenames in os.walk(args.path):
            paths.append(dirpath)
    else:
        paths.append(args.path)

    for path in paths:
        for lasfn in glob.glob(os.path.join(path, args.glob)):
            xlsxfn = lasfn.lower().replace('.las', '.xlsx')
            print('Converting %s -> %s' % (lasfn, xlsxfn))
            try:
                l = las.LASFile(lasfn, ignore_header_errors=args.ignore_header_errors)
                converter = ExcelConverter(l)
                converter.write(xlsxfn)
            except:
                print('Failed to convert file. Error message:\n'
                    + traceback.format_exc())


def get_bulk_parser():
    parser = argparse.ArgumentParser('Convert LAS files to XLSX', 
        formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument('-g', '--glob', default='*.las', help='Match LAS files with this pattern')
    parser.add_argument('-r', '--recursive', action='store_true', help='Recurse through subfolders.', default=False)
    parser.add_argument('-i', '--ignore-header-errors', action='store_true', help='Ignore header section errors.', default=False)
    parser.add_argument('path')
    return parser

if __name__ == '__main__':
    main()
